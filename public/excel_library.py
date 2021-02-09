# -*- coding: utf-8 -*-
"""
 @Time    : 2020/12/22 13:08
 @Author  : GeHH
 功能 : 封装openpyxl库, 提供常用操作excel文件方法
"""
import openpyxl
from openpyxl.styles import PatternFill


class ExcelLibrary:
    """操作xlsx文件, 包括读写"""

    def __init__(self, path):
        self.filepath = path
        self.wbook = openpyxl.load_workbook(self.filepath)
        # 切换默认工作表
        self.sheet = self.wbook.active

    # 改变工作表
    def change_sheet(self, sheet_name):
        self.sheet = self.wbook[sheet_name]

    # 获取工作表名称
    def get_sheetname(self):
        return self.wbook.sheetnames

    # 获取最大行数
    def get_maxrows(self):
        return self.sheet.max_row

    # 获取最大列数
    def get_maxcolumn(self):
        return self.sheet.max_column

    # 提取某一行的值
    def extract_the_row(self, start_column, end_column, row):
        """
        :param startcolumn: 开始列
        :param endcolumn: 结束列
        :param row: 提取行
        :return: 包含指定行项所有值的列表
        """
        row_list = []
        for column in range(start_column, end_column+1):
            row_list.append(self.sheet.cell(row, column).value)
        return row_list

    # 提取某一列的值
    def extract_the_column(self, start_row, end_row, column):
        """
        :param startrow: 开始行
        :param endrow: 结束行
        :param column: 提取列
        :return: 包含指定列所有值的列表
        """
        column_values = []
        for row in range(start_row, end_row+1):
            column_values.append((row, self.sheet.cell(row, column).value))
        return column_values

    # 数据写入单元格
    def write_to_unit(self, row, column, value):
        """单元格插入数据"""
        # 设置单元格颜色

        unit = self.sheet.cell(row, column)
        # 填充背景色
        if value.lower() == 'pass':
            # 绿色7FFF00
            unit.fill = PatternFill("solid", "7FFF00")
        elif value.lower() == 'fail':
            # 红色FF3030
            unit.fill = PatternFill("solid", "FF3030")
        # 写入value
        unit.value = value
        # 保存
        self.save()

    # 保存至文件
    def save(self):
        self.wbook.save(self.filepath)








